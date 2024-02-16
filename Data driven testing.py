import time
from openpyxl.styles import PatternFill
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
# Read data from Excel sheet

workbook = openpyxl.load_workbook('data.xlsx')
sheet = workbook.active
rows_num = sheet.max_row
cols_num = sheet.max_column
driver = webdriver.Chrome()
driver.maximize_window()
driver.get('https://cleartax.in/s/simple-compound-interest-calculator')
for rows in range(2, rows_num+1):
    interest_type = sheet.cell(rows, 1).value
    amount = sheet.cell(rows, 2).value
    rate = sheet.cell(rows, 3).value
    period = sheet.cell(rows, 4).value
    dur = sheet.cell(rows, 5).value
    expected_data = sheet.cell(rows, 6).value
    print(interest_type, amount, rate, period, dur)

    # passing data to application

    type_interest = Select(driver.find_element(By.XPATH, '//select[@id="a"]'))
    type_interest.select_by_visible_text(interest_type)
    driver.find_element(By.XPATH, '//input[@id="c"]').clear()
    driver.find_element(By.XPATH, '//input[@id="c"]').send_keys(amount)
    driver.find_element(By.XPATH, '//input[@id="d"]').clear()
    driver.find_element(By.XPATH, '//input[@id="d"]').send_keys(rate)
    period_type = Select(driver.find_element(By.XPATH, '//select[@id="f"]'))
    period_type.select_by_visible_text(period)
    driver.find_element(By.XPATH, '//input[@id="e"]').clear()
    driver.find_element(By.XPATH, '//input[@id="e"]').send_keys(dur)

    final_value = driver.find_element(By.XPATH, '(//div[@class="output"])[3]/span').text.replace('₹ ', '').\
        replace(',', '')

    # validating the data
    
    if float(final_value) == float(expected_data):
        sheet.cell(rows, 7).value = "Passed"
        sheet.cell(rows, 7).fill = PatternFill(patternType='solid',
                            fgColor='00CC00')
    else:
        sheet.cell(rows, 7).value = "Failed"
        sheet.cell(rows, 7).fill = PatternFill(patternType='solid',
                                               fgColor='C64747')
    workbook.save('data.xlsx')
    time.sleep(5)

driver.quit()


